from __future__ import annotations

import csv
import copy
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "Voeding prijzen en afbeeldingen" / "inkoopmatrix_21_beslisboom_urls.xlsx"
BROK_CSV_PATH = ROOT / "Voeding prijzen en afbeeldingen" / "CSV files" / "brok-voerlijst-2026-03-11.csv"
NATVOER_CSV_PATH = ROOT / "Voeding prijzen en afbeeldingen" / "CSV files" / "natvoer-voerlijst-2026-03-11.csv"
IMAGE_DIR = ROOT / "Voeding prijzen en afbeeldingen" / "Afbeeldingen 21 SKU's"


KITTEN_MATRIX_ROWS = [
    {
        "Matrix_ID": "MX-22",
        "Type": "Brok",
        "Profiel": "Kitten",
        "Budget/Lane": "Laag",
        "Merk": "Josera",
        "Lijn": "Kitten – Graanvrij",
        "Smaak": "Gevogelte",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 0.44,
        "CSV prijs per maand (€)": 13.20,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Vanaf €4,99 (400 g)",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_droog/josera/kattenvoer/986866",
        "Conditiedekking": "kitten groei; dagelijkse basis; kleinere brok",
        "Fallback alternatief": "—",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
    },
    {
        "Matrix_ID": "MX-23",
        "Type": "Brok",
        "Profiel": "Kitten",
        "Budget/Lane": "Midden",
        "Merk": "Carnilove",
        "Lijn": "Kitten – Salmon & Turkey",
        "Smaak": "Zalm en kalkoen",
        "Geselecteerde leverancier": "Poezenparadijs",
        "CSV prijs per dag (€)": 0.62,
        "CSV prijs per maand (€)": 18.60,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijsrange zichtbaar op leverancier",
        "Product URL": "https://poezenparadijs.nl/product/carnilove-salmon-turkey-kittenvoer/",
        "Conditiedekking": "kitten groei; hoge dierlijke eiwitbasis; premium middenklasse",
        "Fallback alternatief": "Josera Kitten Graanvrij",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
    },
    {
        "Matrix_ID": "MX-24",
        "Type": "Brok",
        "Profiel": "Kitten",
        "Budget/Lane": "Hoog",
        "Merk": "Farmina",
        "Lijn": "N&D Prime – Kitten",
        "Smaak": "Kip en granaatappel",
        "Geselecteerde leverancier": "Poezenparadijs",
        "CSV prijs per dag (€)": 1.08,
        "CSV prijs per maand (€)": 32.40,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijsrange zichtbaar op leverancier",
        "Product URL": "https://poezenparadijs.nl/product/farmina-nd-prime-kip-granaatappel-voor-kittens/",
        "Conditiedekking": "kitten groei; rijke eiwitbron; luxe ontwikkellijn",
        "Fallback alternatief": "Carnilove Kitten Salmon & Turkey",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
    },
    {
        "Matrix_ID": "MX-25",
        "Type": "Natvoer",
        "Profiel": "Kitten",
        "Budget/Lane": "Laag",
        "Merk": "Bozita",
        "Lijn": "Hapjes in saus – Kitten",
        "Smaak": "Kip",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 1.48,
        "CSV prijs per maand (€)": 44.40,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijs zichtbaar op productpagina",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/bozita/bozita_saus/1412223",
        "Conditiedekking": "kitten natvoer; hydratatie; laag budget",
        "Fallback alternatief": "Feringa Classic Meat Menu Kitten",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
    },
    {
        "Matrix_ID": "MX-26",
        "Type": "Natvoer",
        "Profiel": "Kitten",
        "Budget/Lane": "Midden",
        "Merk": "Feringa",
        "Lijn": "Classic Meat Menu – Kitten",
        "Smaak": "Kip en zalm",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 1.64,
        "CSV prijs per maand (€)": 49.20,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijs zichtbaar op productpagina",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/feringa/feringa_kitten/490287",
        "Conditiedekking": "kitten natvoer; groei; midden budget",
        "Fallback alternatief": "Bozita Hapjes in saus – Kitten",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
    },
    {
        "Matrix_ID": "MX-27",
        "Type": "Natvoer",
        "Profiel": "Kitten",
        "Budget/Lane": "Hoog",
        "Merk": "Catz Finefood",
        "Lijn": "Kitten",
        "Smaak": "Gevogelte",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 1.98,
        "CSV prijs per maand (€)": 59.40,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijs zichtbaar op productpagina",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/catz_finefood/kitten/1389131",
        "Conditiedekking": "kitten natvoer; premium groei; hoge acceptatie",
        "Fallback alternatief": "Feringa Classic Meat Menu – Kitten",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
    },
]

KITTEN_SKU_ROWS = [
    {
        "Unique_SKU": "SKU-20",
        "Matrix_ID": "MX-22",
        "Type": "Brok",
        "Profiel": "Kitten",
        "Budget/Lane": "Laag",
        "Merk": "Josera",
        "Lijn": "Kitten – Graanvrij",
        "Smaak": "Gevogelte",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 0.44,
        "CSV prijs per maand (€)": 13.20,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Vanaf €4,99 (400 g)",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_droog/josera/kattenvoer/986866",
        "Conditiedekking": "kitten groei; dagelijkse basis; kleinere brok",
        "Fallback alternatief": "—",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
        "Unique_SKU2": "",
    },
    {
        "Unique_SKU": "SKU-21",
        "Matrix_ID": "MX-23",
        "Type": "Brok",
        "Profiel": "Kitten",
        "Budget/Lane": "Midden",
        "Merk": "Carnilove",
        "Lijn": "Kitten – Salmon & Turkey",
        "Smaak": "Zalm en kalkoen",
        "Geselecteerde leverancier": "Poezenparadijs",
        "CSV prijs per dag (€)": 0.62,
        "CSV prijs per maand (€)": 18.60,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijsrange zichtbaar op leverancier",
        "Product URL": "https://poezenparadijs.nl/product/carnilove-salmon-turkey-kittenvoer/",
        "Conditiedekking": "kitten groei; hoge dierlijke eiwitbasis; premium middenklasse",
        "Fallback alternatief": "Josera Kitten Graanvrij",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
        "Unique_SKU2": "",
    },
    {
        "Unique_SKU": "SKU-22",
        "Matrix_ID": "MX-24",
        "Type": "Brok",
        "Profiel": "Kitten",
        "Budget/Lane": "Hoog",
        "Merk": "Farmina",
        "Lijn": "N&D Prime – Kitten",
        "Smaak": "Kip en granaatappel",
        "Geselecteerde leverancier": "Poezenparadijs",
        "CSV prijs per dag (€)": 1.08,
        "CSV prijs per maand (€)": 32.40,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijsrange zichtbaar op leverancier",
        "Product URL": "https://poezenparadijs.nl/product/farmina-nd-prime-kip-granaatappel-voor-kittens/",
        "Conditiedekking": "kitten groei; rijke eiwitbron; luxe ontwikkellijn",
        "Fallback alternatief": "Carnilove Kitten Salmon & Turkey",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
        "Unique_SKU2": "",
    },
    {
        "Unique_SKU": "SKU-23",
        "Matrix_ID": "MX-25",
        "Type": "Natvoer",
        "Profiel": "Kitten",
        "Budget/Lane": "Laag",
        "Merk": "Bozita",
        "Lijn": "Hapjes in saus – Kitten",
        "Smaak": "Kip",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 1.48,
        "CSV prijs per maand (€)": 44.40,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijs zichtbaar op productpagina",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/bozita/bozita_saus/1412223",
        "Conditiedekking": "kitten natvoer; hydratatie; laag budget",
        "Fallback alternatief": "Feringa Classic Meat Menu Kitten",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
        "Unique_SKU2": "",
    },
    {
        "Unique_SKU": "SKU-24",
        "Matrix_ID": "MX-26",
        "Type": "Natvoer",
        "Profiel": "Kitten",
        "Budget/Lane": "Midden",
        "Merk": "Feringa",
        "Lijn": "Classic Meat Menu – Kitten",
        "Smaak": "Kip en zalm",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 1.64,
        "CSV prijs per maand (€)": 49.20,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijs zichtbaar op productpagina",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/feringa/feringa_kitten/490287",
        "Conditiedekking": "kitten natvoer; groei; midden budget",
        "Fallback alternatief": "Bozita Hapjes in saus – Kitten",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
        "Unique_SKU2": "",
    },
    {
        "Unique_SKU": "SKU-25",
        "Matrix_ID": "MX-27",
        "Type": "Natvoer",
        "Profiel": "Kitten",
        "Budget/Lane": "Hoog",
        "Merk": "Catz Finefood",
        "Lijn": "Kitten",
        "Smaak": "Gevogelte",
        "Geselecteerde leverancier": "Zooplus",
        "CSV prijs per dag (€)": 1.98,
        "CSV prijs per maand (€)": 59.40,
        "Leverbaar?": "Ja",
        "Website prijs op leverancier": "Prijs zichtbaar op productpagina",
        "Product URL": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/catz_finefood/kitten/1389131",
        "Conditiedekking": "kitten natvoer; premium groei; hoge acceptatie",
        "Fallback alternatief": "Feringa Classic Meat Menu – Kitten",
        "Opmerking": "Kitten-lane toegevoegd op basis van actuele productpagina, analyse en lokale app-uitbreiding.",
        "Unique_SKU2": "",
    },
]

BROK_CSV_ROWS = [
    {
        "web_scraper_order": "kitten-mx-22",
        "web_scraper_start_url": "https://www.zooplus.nl/shop/katten/kattenvoer_droog/josera/kattenvoer/986866",
        "pagination": "",
        "vet": "22%",
        "data": "10%",
        "kcal": "4.19",
        "vezels": "2.0%",
        "lijn": "Kitten – Graanvrij",
        "€dag": "€ 0.44",
        "fosfor": "1.05%",
        "eiwit": "36%",
        "ca:p": "1.3",
        "shop": "Zooplus",
        "merk": "Josera",
        "merk2": "Josera",
        "intact": "1",
        "castraat": "1",
        "ingr3": "1",
        "sensitive": "2",
        "smaak": "Gevogelte",
        "graanvrij2": "J",
        "bio2": "N",
        "large": "N",
        "image": "https://media.zooplus.com/bilder/3/400/_pla_foodforplanet_josera_kitten_getreidefrei_hs_01_3.jpg",
        "ingr": "1",
        "ingr2": "",
        "bio": "N",
        "graanvrij": "J",
        "fosfor2": "",
    },
    {
        "web_scraper_order": "kitten-mx-23",
        "web_scraper_start_url": "https://poezenparadijs.nl/product/carnilove-salmon-turkey-kittenvoer/",
        "pagination": "",
        "vet": "20%",
        "data": "10%",
        "kcal": "4.12",
        "vezels": "2.0%",
        "lijn": "Kitten – Salmon & Turkey",
        "€dag": "€ 0.62",
        "fosfor": "1.0%",
        "eiwit": "40%",
        "ca:p": "1.3",
        "shop": "Poezenparadijs",
        "merk": "Carnilove",
        "merk2": "Carnilove",
        "intact": "1",
        "castraat": "1",
        "ingr3": "1",
        "sensitive": "2",
        "smaak": "Zalm en kalkoen",
        "graanvrij2": "J",
        "bio2": "N",
        "large": "N",
        "image": "https://poezenparadijs.nl/wp-content/uploads/2021/03/Carnilove-kittenvoer.jpg",
        "ingr": "1",
        "ingr2": "",
        "bio": "N",
        "graanvrij": "J",
        "fosfor2": "",
    },
    {
        "web_scraper_order": "kitten-mx-24",
        "web_scraper_start_url": "https://poezenparadijs.nl/product/farmina-nd-prime-kip-granaatappel-voor-kittens/",
        "pagination": "",
        "vet": "20%",
        "data": "8%",
        "kcal": "4.29",
        "vezels": "1.8%",
        "lijn": "N&D Prime – Kitten",
        "€dag": "€ 1.08",
        "fosfor": "0.9%",
        "eiwit": "44%",
        "ca:p": "1.2",
        "shop": "Poezenparadijs",
        "merk": "Farmina",
        "merk2": "Farmina",
        "intact": "1",
        "castraat": "1",
        "ingr3": "1",
        "sensitive": "2",
        "smaak": "Kip en granaatappel",
        "graanvrij2": "J",
        "bio2": "N",
        "large": "N",
        "image": "https://poezenparadijs.nl/wp-content/uploads/2021/09/Farmina-ND-Chicken-pomegranate-kitten.png",
        "ingr": "1",
        "ingr2": "",
        "bio": "N",
        "graanvrij": "J",
        "fosfor2": "",
    },
]

NATVOER_CSV_ROWS = [
    {
        "web_scraper_order": "kitten-mx-25",
        "web_scraper_start_url": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/bozita/bozita_saus/1412223",
        "pagination": "",
        "smaak": "Kip",
        "€dag": "€ 1.48",
        "eiwit": "8.5%",
        "vet": "5.5%",
        "data": "82%",
        "kcal": "0.86",
        "lijn": "Hapjes in saus – Kitten",
        "vezels": "0.5%",
        "textuur": "Saus",
        "shop": "Zooplus",
        "merk": "Bozita",
        "merk2": "Bozita",
        "fosfor": "",
        "ca:p": "",
        "intact": "1",
        "castraat": "1",
        "sensitive2": "2",
        "graanvrij2": "N",
        "bio2": "N",
        "ingr2": "1",
        "fosfor2": "",
        "sensitive": "",
        "graanvrij": "",
        "bio": "",
        "ingr": "",
    },
    {
        "web_scraper_order": "kitten-mx-26",
        "web_scraper_start_url": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/feringa/feringa_kitten/490287",
        "pagination": "",
        "smaak": "Kip en zalm",
        "€dag": "€ 1.64",
        "eiwit": "11.1%",
        "vet": "6.6%",
        "data": "79%",
        "kcal": "1.03",
        "lijn": "Classic Meat Menu – Kitten",
        "vezels": "0.3%",
        "textuur": "Paté",
        "shop": "Zooplus",
        "merk": "Feringa",
        "merk2": "Feringa",
        "fosfor": "0.19%",
        "ca:p": "",
        "intact": "1",
        "castraat": "1",
        "sensitive2": "2",
        "graanvrij2": "J",
        "bio2": "N",
        "ingr2": "1",
        "fosfor2": "",
        "sensitive": "",
        "graanvrij": "",
        "bio": "",
        "ingr": "",
    },
    {
        "web_scraper_order": "kitten-mx-27",
        "web_scraper_start_url": "https://www.zooplus.nl/shop/katten/kattenvoer_blik/catz_finefood/kitten/1389131",
        "pagination": "",
        "smaak": "Gevogelte",
        "€dag": "€ 1.98",
        "eiwit": "11.0%",
        "vet": "5.9%",
        "data": "80%",
        "kcal": "0.95",
        "lijn": "Kitten",
        "vezels": "0.5%",
        "textuur": "Paté",
        "shop": "Zooplus",
        "merk": "Catz Finefood",
        "merk2": "Catz Finefood",
        "fosfor": "",
        "ca:p": "",
        "intact": "1",
        "castraat": "1",
        "sensitive2": "2",
        "graanvrij2": "J",
        "bio2": "N",
        "ingr2": "1",
        "fosfor2": "",
        "sensitive": "",
        "graanvrij": "",
        "bio": "",
        "ingr": "",
    },
]

IMAGE_DOWNLOADS = [
    (
        "MX-22_Josera_Kitten_Graanvrij_Gevogelte.jpg",
        "https://media.zooplus.com/bilder/3/400/_pla_foodforplanet_josera_kitten_getreidefrei_hs_01_3.jpg",
    ),
    (
        "MX-23_Carnilove_Kitten_Salmon_Turkey_Zalm_en_kalkoen.jpg",
        "https://poezenparadijs.nl/wp-content/uploads/2021/03/Carnilove-kittenvoer.jpg",
    ),
    (
        "MX-24_Farmina_ND_Prime_Kitten_Kip_en_granaatappel.png",
        "https://poezenparadijs.nl/wp-content/uploads/2021/09/Farmina-ND-Chicken-pomegranate-kitten.png",
    ),
    (
        "MX-25_Bozita_Hapjes_in_saus_Kitten_Kip.jpg",
        "https://media.zooplus.com/bilder/7/400/332210_bozita_happchen_sosse_kitten_huhn_12x85g_hs_02_7.jpg",
    ),
    (
        "MX-26_Feringa_Classic_Meat_Menu_Kitten_Kip_en_zalm.jpg",
        "https://media.zooplus.com/bilder/8/400/2024_old_new_feringa_classic_kitten_chicken_salmon_200g_1_8.jpg",
    ),
    (
        "MX-27_Catz_Finefood_Kitten_Gevogelte.jpg",
        "https://media.zooplus.com/bilder/9/400/297096_pla_petsnature_catzfinefood_kitten_gefluegel_85g_hs_01_9.jpg",
    ),
]


def update_csv(path: Path, id_key: str, rows: list[dict]) -> None:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        existing_rows = list(reader)
        fieldnames = reader.fieldnames or []

    row_map = {row[id_key]: row for row in existing_rows if row.get(id_key)}
    for row in rows:
        row_map[row[id_key]] = row

    preserved = [row for row in existing_rows if row.get(id_key) not in row_map or row.get(id_key) not in {r[id_key] for r in rows}]
    merged = preserved + [row_map[row[id_key]] for row in rows]

    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(merged)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target._style = copy.copy(source._style)
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
        target.number_format = source.number_format


def update_sheet_by_key(ws, key_name: str, rows: list[dict], style_source_row: int) -> None:
    headers = [cell.value for cell in ws[1]]
    key_col = headers.index(key_name) + 1
    row_index_by_key = {}
    for row_idx in range(2, ws.max_row + 1):
        key = ws.cell(row_idx, key_col).value
        if key:
            row_index_by_key[str(key)] = row_idx

    for row in rows:
        key = str(row[key_name])
        if key in row_index_by_key:
            target_row = row_index_by_key[key]
        else:
            target_row = ws.max_row + 1
            copy_row_style(ws, style_source_row, target_row, len(headers))

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(target_row, col_idx).value = row.get(header, "")


def update_workbook() -> None:
    wb = load_workbook(WORKBOOK_PATH)

    update_sheet_by_key(wb["Inkoopmatrix_21"], "Matrix_ID", KITTEN_MATRIX_ROWS, style_source_row=22)
    update_sheet_by_key(wb["SKU_Unique"], "Matrix_ID", KITTEN_SKU_ROWS, style_source_row=12)

    beslisboom = wb["Beslisboom"]
    beslisboom["C2"] = "Kitten-lane: brok MX-22 / MX-23 / MX-24, nat MX-25 / MX-26 / MX-27"
    beslisboom["D2"] = "Ga naar stap 2"
    beslisboom["E2"] = "Kitten is nu als aparte lane opgenomen, zodat groei, energiedichtheid en kleinere porties apart beoordeeld worden."

    summary = wb["Samenvatting"]
    summary["B4"] = 27
    summary["E4"] = "27 rijen: 18 basis + 3 specialist + 6 kitten"
    summary["B5"] = 25
    summary["E5"] = "Met de kitten-lane erbij zijn er nu 25 unieke producten"
    summary["B6"] = 20
    summary["E6"] = "Zooplus blijft hoofdleverancier met ook de kitten-natlijn"
    summary["B7"] = 5
    summary["E7"] = "Poezenparadijs blijft specialistische leverancier voor premium kitten- en specialistische brok"
    summary["B8"] = 0.8
    summary["E8"] = "Ondanks de kitten-uitbreiding blijft de leveranciersmix compact"

    wb.save(WORKBOOK_PATH)


def download_images() -> None:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": "Mozilla/5.0"}
    for filename, url in IMAGE_DOWNLOADS:
        target = IMAGE_DIR / filename
        request = Request(url, headers=headers)
        with urlopen(request) as response:
            target.write_bytes(response.read())


def main() -> None:
    update_workbook()
    update_csv(BROK_CSV_PATH, "web_scraper_order", BROK_CSV_ROWS)
    update_csv(NATVOER_CSV_PATH, "web_scraper_order", NATVOER_CSV_ROWS)
    download_images()
    print("Kitten lane added to workbook, CSV files, and image directory.")


if __name__ == "__main__":
    main()
